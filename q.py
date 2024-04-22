from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions

# Создание объекта опций
options = Options()
# Инициализация веб-драйвера с использованием объекта опций
driver = webdriver.Chrome(options=options, executable_path='chromedriver.exe')
driver.get('https://www.neftegaz-expo.ru/ru/visitors/mm/?auto=mm&step=step1')
time.sleep(5)
# ПРОЦЕСС РЕГИСТРАЦИИ
# Присваиваем переменную для поля логина
xpath_login = '//*[@id="EmailText"]'
# Записываем поиск этого элемента
input_element = driver.find_element(By.XPATH, xpath_login)
# print("111")
# Производим запись (ВВОД ЛОГИНА)
input_element.send_keys("ponoxaw267@adstam.com")
# input_element.send_keys("e.maksimova@e-shkaf.su")
# Присваиваем переменную для кнопки
xpath_login_button = '//*[@id="expodatakrat"]/div/div/div[2]/div/div[2]/button'
# Ищем кнопку
input_element = driver.find_element(By.XPATH, xpath_login_button)
# Жмем на кнопку
input_element.click()
# Время ожидания
time.sleep(5)
# Присваиваем переменную для поля пароля
xpath_password = '//*[@id="Password"]'
# Записываем поиск этого элемента
input_element = driver.find_element(By.XPATH, xpath_password)
# Производим запись (ВВОД ПАРОЛЯ)
# input_element.send_keys("231287")
input_element.send_keys("3452394789")
# Присваиваем переменную для кнопки
xpath_password_button = '//*[@id="expodatakrat"]/div/div/div[2]/div[1]/div[3]/button'
# Ищем кнопку
input_element = driver.find_element(By.XPATH, xpath_password_button)
# Жмем на кнопку
input_element.click()
# Время ожидания 3452394789
time.sleep(5)
# Обновление страницы для перехода на другой язык
# driver.get("https://www.neftegaz-expo.ru/ru/visitors/mm/?cabinet&activityId=1466&visitorId=3491128#%"+"2FMatchMaking%"+"2FCompany%"+"2FList%"+"3Fpage=97")
driver.refresh()
objects = []
from selenium.webdriver.common.action_chains import ActionChains
actions = ActionChains(driver)

while True:
        try:
            list_items_block = driver.find_element(By.XPATH, '//*[@id="widgetInner"]/div/div[1]')
            n=0
            items = list_items_block.find_elements(By.CLASS_NAME, "mm_company_list_item")
            for item in items:
                actions.move_to_element(item).perform()
                name = item.find_element(By.CLASS_NAME, "mm_company_list_link").text
                tags = item.find_element(By.CLASS_NAME, "mm_company_list_item_rubric_settings").text
                addr = item.find_element(By.CLASS_NAME, "mm_company_address").text
                img = item.find_element(By.CLASS_NAME, "mm_company_list_item_logo").find_element(By.TAG_NAME, "img").get_attribute("src")
                n+=1
                # Убираем полученное и получаем оставшееся
                deskr = item.text.replace(name,"").replace(tags,"").replace(addr,"").replace(img,"").replace("Подробнее о компании","").replace("Назначить встречу","")
                print(n)
                new_item = [name, tags, addr,deskr, img]
                objects.append(new_item)
            # перейти к слеующей странице
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            nums = driver.find_elements(By.CLASS_NAME, 'bottom-element')
            if len(nums) != 0:
                # print(new_item)
                print(' '.join(map(lambda x: x.text, nums)))
                for obj in nums:
                    if "[" in obj.text:
                         index = nums.index(obj)
                         next = nums[index+1].find_element(By.TAG_NAME, "a")
                         print(next.text+" - next")
                         next.click()
                         driver.refresh()
                         time.sleep(2)
            else:
                print("wait нумерация внизу не найдена")
                time.sleep(1)
                print("__________________________________________________________________")
        except IndexError:
             print("error index")
             break
        except Exception as e:
            print(e)
            print("OUT Exception")
            time.sleep(1)
print(objects)
# я на ленте
# скрол вниз
# беру все карты
# переходу на следующую если есть

from openpyxl import Workbook

def create_excel_file(filename, headers, data):
    wb = Workbook()
    ws = wb.active
    
    # Задаем заголовки столбцов
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    
    # Записываем данные в определенные ячейки
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, cell_data in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell_data)
    
    # Сохраняем файл Excel
    wb.save(filename)
    print(f"Файл '{filename}' успешно создан или перезаписан.")

# Пример использования
filename = "test2.xlsx"
headers = ["Имя", "Тэги", "Адрес", "Описание", "Лого"]
data = [
    ["Иван", 25, "ivan@example.com"],
    ["Мария", 30, "maria@example.com"],
    ["Петр", 35, "petr@example.com"]
]

create_excel_file(filename, headers, objects)